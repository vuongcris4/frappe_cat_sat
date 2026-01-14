import frappe
from frappe import _
from frappe.utils.file_manager import save_file
import json
import io

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None


@frappe.whitelist()
def download_template():
    """Generate and return Excel template for Cutting Specification import"""
    if not openpyxl:
        frappe.throw("Thư viện 'openpyxl' chưa được cài đặt. Vui lòng cài đặt: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cutting Specification"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        ("A", "spec_name", "Tên định mức"),
        ("B", "piece_code", "Mã mảnh"),
        ("C", "piece_name", "Tên mảnh"),
        ("D", "piece_qty", "SL mảnh/TP"),
        ("E", "steel_profile", "Loại sắt"),
        ("F", "segment_name", "Tên đoạn sắt"),
        ("G", "length_mm", "Chiều dài (mm)"),
        ("H", "qty_per_piece", "SL đoạn/mảnh"),
        ("I", "punch_hole_qty", "Số lỗ dập"),
        ("J", "rivet_hole_qty", "Số lỗ tán"),
        ("K", "bend_type", "Uốn"),
        ("L", "note", "Ghi chú"),
    ]
    
    for col, field_name, label in headers:
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
        # Field name row (row 2)
        field_cell = ws[f"{col}2"]
        field_cell.value = field_name
        field_cell.font = Font(italic=True, color="808080")
        field_cell.border = thin_border
    
    # Sample data
    sample_data = [
        ["i3", "I31.27.1", "Ghế góc", 1, "", "", "", "", "", "", "", ""],
        ["i3", "I31.27.1.1", "Mê ngồi", 1, "", "", "", "", "", "", "", ""],
        ["i3", "I31.27.1.1.1", "", "", "V15", "Thanh ngang", 1162.2, 1, 2, 2, "Không", "2 dập, 2 tán"],
        ["i3", "I31.27.1.1.2", "", "", "V15", "Thanh dọc", 2375.8, 1, 2, 1, "Không", "2 dập, 1 tán"],
        ["i3", "I31.27.1.1.3", "", "", "FI10", "Ống tròn", 270, 1, 0, 0, "Không", ""],
        ["i3", "I31.27.1.1.4", "", "", "FI10", "Chân ghế", 192, 4, 0, 0, "Không", ""],
        ["i3", "I31.27.1.2", "Tựa góc nhỏ", 1, "", "", "", "", "", "", "", ""],
        ["i3", "I31.27.1.2.1", "", "", "V15", "Tựa lưng", 565, 1, 2, 1, "Không", "2 tán, 1 tựa"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=3):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.value = value
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = [15, 15, 20, 12, 12, 20, 15, 12, 10, 10, 12, 25]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width
    
    # Instructions sheet
    ws_help = wb.create_sheet("Hướng dẫn")
    instructions = [
        ("Hướng dẫn Import Cutting Specification", ""),
        ("", ""),
        ("1. Cột spec_name", "Tên định mức cắt (bắt buộc cho dòng đầu tiên của mỗi spec)"),
        ("2. Cột piece_code", "Mã mảnh phân cấp (VD: I31.27.1.1.1)"),
        ("3. Cột piece_name", "Tên mảnh (nếu có)"),
        ("4. Cột piece_qty", "Số lượng mảnh / thành phẩm"),
        ("5. Cột steel_profile", "Loại sắt (phải tồn tại trong Steel Profile)"),
        ("6. Cột segment_name", "Tên đoạn sắt"),
        ("7. Cột length_mm", "Chiều dài đoạn sắt (mm)"),
        ("8. Cột qty_per_piece", "Số đoạn giống nhau trong 1 mảnh"),
        ("9. Cột punch_hole_qty", "Số lỗ dập"),
        ("10. Cột rivet_hole_qty", "Số lỗ tán"),
        ("11. Cột bend_type", "Loại uốn: Không / Uốn 1 đầu / Uốn 2 đầu / Uốn cong"),
        ("12. Cột note", "Ghi chú gia công"),
        ("", ""),
        ("LƯU Ý:", ""),
        ("- Dòng 1 là tiêu đề, dòng 2 là tên field", ""),
        ("- Bắt đầu nhập dữ liệu từ dòng 3", ""),
        ("- Mỗi dòng có piece_name sẽ tạo mảnh mới", ""),
        ("- Mỗi dòng có steel_profile sẽ tạo chi tiết sắt", ""),
    ]
    
    for row_idx, (col1, col2) in enumerate(instructions, start=1):
        ws_help.cell(row=row_idx, column=1).value = col1
        ws_help.cell(row=row_idx, column=2).value = col2
        if row_idx == 1:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    ws_help.column_dimensions['A'].width = 25
    ws_help.column_dimensions['B'].width = 50
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as file download
    frappe.response['filename'] = 'cutting_specification_template.xlsx'
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def import_from_excel(file_url):
    """Import Cutting Specification from uploaded Excel file"""
    if not openpyxl:
        frappe.throw("Thư viện 'openpyxl' chưa được cài đặt.")
    
    # Get file content
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Skip header rows (row 1 = labels, row 2 = field names)
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    
    if not rows:
        frappe.throw("File Excel không có dữ liệu.")
    
    # Group by spec_name
    specs = {}
    current_spec = None
    current_piece = None
    
    for row in rows:
        spec_name = row[0] or current_spec
        piece_code = row[1]
        piece_name = row[2]
        piece_qty = row[3]
        steel_profile = row[4]
        segment_name = row[5]
        length_mm = row[6]
        qty_per_piece = row[7]
        punch_hole_qty = row[8]
        rivet_hole_qty = row[9]
        bend_type = row[10]
        note = row[11] if len(row) > 11 else ""
        
        if not spec_name:
            continue
        
        current_spec = spec_name
        
        if spec_name not in specs:
            specs[spec_name] = {
                "pieces": {},
                "details": []
            }
        
        # If row has piece_name, add/update piece
        if piece_name:
            current_piece = piece_name
            if piece_name not in specs[spec_name]["pieces"]:
                specs[spec_name]["pieces"][piece_name] = {
                    "piece_code": piece_code,
                    "piece_name": piece_name,
                    "piece_qty": piece_qty or 1
                }
        
        # If row has steel_profile and length, add detail
        if steel_profile and length_mm:
            # Validate steel_profile exists
            if not frappe.db.exists("Steel Profile", steel_profile):
                frappe.throw(f"Steel Profile '{steel_profile}' không tồn tại. Vui lòng tạo trước.")
            
            specs[spec_name]["details"].append({
                "piece_name": current_piece or piece_code,
                "steel_profile": steel_profile,
                "segment_name": segment_name or piece_code,
                "length_mm": float(length_mm),
                "qty_segment_per_piece": int(qty_per_piece or 1),
                "punch_hole_qty": int(punch_hole_qty or 0),
                "rivet_hole_qty": int(rivet_hole_qty or 0),
                "bend_type": bend_type or "Không",
                "note": note or ""
            })
    
    # Create Cutting Specifications
    created = []
    for spec_name, data in specs.items():
        # Check if exists
        if frappe.db.exists("Cutting Specification", {"spec_name": spec_name}):
            frappe.throw(f"Cutting Specification '{spec_name}' đã tồn tại.")
        
        doc = frappe.new_doc("Cutting Specification")
        doc.spec_name = spec_name
        
        # Add pieces
        for piece_data in data["pieces"].values():
            doc.append("pieces", piece_data)
        
        # Add details
        for detail in data["details"]:
            doc.append("details", detail)
        
        doc.insert()
        created.append(spec_name)
    
    frappe.db.commit()
    
    return {
        "message": f"Đã import thành công {len(created)} Cutting Specification",
        "created": created
    }
