"""
Import script for Định Mức Vật Tư (Material Specifications)
All measurements converted from cm to mm (*10)
"""
import frappe
import openpyxl
import csv
import os
import re
import glob

INPUT_DIR = "/home/trand/frappe/sites/erp.dongnama.app/workspace/frappe-bench/input_data"
PHOISAT_DIR = f"{INPUT_DIR}/PhoiSat"

# Steel profile name normalization mapping
STEEL_PROFILE_MAPPING = {
    # V profiles
    'v10': 'V10', 'v 10': 'V10', 'sắt v10': 'V10', 'v10 6 zem': 'V10', 'v10 6zem': 'V10', 'v10*6 zem': 'V10', 'v10*6zem': 'V10',
    'v12': 'V12', 'v 12': 'V12', 'sắt v12': 'V12', 'v12 6 zem': 'V12', 'v12 6zem': 'V12', 'v12*6 zem': 'V12',
    'v14': 'V14', 'v 14': 'V14', 'v14*6 zem': 'V14',
    'v15': 'V15', 'v 15': 'V15', 'sắt v15': 'V15', 'v15 6 zem': 'V15', 'v15 6zem': 'V15', 'v15 7zem': 'V15', 'v15 7 zem': 'V15', 'v15*6 zem': 'V15', 'v15*7 zem': 'V15',
    'v16': 'V16', 'v 16': 'V16',
    'v18': 'V18', 'v 18': 'V18', 'v18 6 zem': 'V18', 'v18 6zem': 'V18',
    'v20': 'V20', 'v 20': 'V20', 'sắt v20': 'V20', 'v20 6 zem': 'V20', 'v20 8zem': 'V20', 'v20*6zem': 'V20', 'v20* 0.6 zem': 'V20', 'v20* 8 zem': 'V20', 'săt v20 6zem': 'V20',
    'v25': 'V25', 'v 25': 'V25', 'sắt v25': 'V25', 'v25 6 zem': 'V25', 'v25*6 zem': 'V25',
    'v30': 'V30', 'v 30': 'V30', 'sắt v30 6 zem': 'V30',
    'v40': 'V40', 'v 40': 'V40',
    'v50': 'V50', 'v 50': 'V50',
    # Fi profiles  
    'fi 4': 'Fi4', 'fi4': 'Fi4', 'sắt fi 4': 'Fi4', 'thép fi 4': 'Fi4',
    'fi 6': 'Fi6', 'fi6': 'Fi6', 'sắt fi 6': 'Fi6', 'sắt fi6': 'Fi6',
    'fi 8': 'Fi8', 'fi8': 'Fi8', 'sắt fi 8': 'Fi8', 'sắt fi 08': 'Fi8',
    'fi 10': 'FI10', 'fi10': 'FI10', 'sắt fi 10': 'FI10', 'sắt fi10': 'FI10', 'sắt fi 10 6 zem': 'FI10', 'fi 10 6zem': 'FI10',
    'fi 16': 'Fi16', 'fi16': 'Fi16',
    'fi 19': 'FI19', 'fi19': 'FI19', 'sắt fi 19': 'FI19', 'sắt fi 19 6 zem': 'FI19',
    'fi 21': 'Fi21', 'fi21': 'Fi21',
    # H profiles
    'h10-20': 'H10-20', 'h10*20': 'H10-20', 'sắt h10-20': 'H10-20', 'sắt h10*20 6 zem': 'H10-20', 'săt h10*20 6zem': 'H10-20', 'h10-20 6 zem': 'H10-20', 'h10-20 6zem': 'H10-20',
    'h13-26': 'H13-26', 'h13*26': 'H13-26',
    'h15-35': 'H15-35', 'h15*35': 'H15-35',
    'h20-40': 'H20-40', 'h20*40': 'H20-40',
    'h25-50': 'H25-50', 'h25*50': 'H25-50',
}

def normalize_steel_profile(name):
    """Normalize steel profile name to standard format"""
    if not name:
        return None
    
    # Clean the name
    clean = name.lower().strip()
    # Remove processing notes in parentheses but keep them for reference
    base_name = re.sub(r'\s*\([^)]*\)', '', clean).strip()
    
    # Try direct mapping
    if base_name in STEEL_PROFILE_MAPPING:
        return STEEL_PROFILE_MAPPING[base_name]
    
    # Try to extract profile pattern
    patterns = [
        (r'^v\s*(\d+)', lambda m: f'V{m.group(1)}'),
        (r'^fi\s*(\d+)', lambda m: f'Fi{m.group(1)}'),
        (r'^h(\d+)[-*](\d+)', lambda m: f'H{m.group(1)}-{m.group(2)}'),
    ]
    for pattern, formatter in patterns:
        match = re.match(pattern, base_name)
        if match:
            return formatter(match)
    
    # Return original if no match
    return name.strip()

def create_missing_steel_profiles():
    """Create Steel Profiles that don't exist yet"""
    profiles_to_create = [
        'V14', 'V16', 'V18', 'V25', 'V30', 'V40',
        'Fi4', 'Fi8',
        'H13-26', 'H15-35', 'H20-40'
    ]
    
    created = []
    for profile_name in profiles_to_create:
        if not frappe.db.exists('Steel Profile', profile_name):
            doc = frappe.get_doc({
                'doctype': 'Steel Profile',
                'profile_name': profile_name
            })
            doc.insert(ignore_permissions=True)
            created.append(profile_name)
            print(f"✓ Created Steel Profile: {profile_name}")
        else:
            print(f"- Steel Profile exists: {profile_name}")
    
    frappe.db.commit()
    return created

def parse_phoisat_file(filepath):
    """Parse a PhoiSat file and extract cutting details"""
    details = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_piece = None
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            vals = [cell.value for cell in row[:10]]
            
            # Skip empty rows
            if not any(vals):
                continue
            
            # Check for piece header (e.g., "I3GỖ.1.1 Khung tựa")
            first_val = str(vals[0]).strip() if vals[0] else ''
            
            # Look for segment rows with code pattern like I3GỖ.1.1.1
            if vals[0] and vals[1] and vals[2] and vals[3]:
                segment_code = str(vals[0]).strip()
                steel_name = str(vals[1]).strip()
                unit = str(vals[2]).strip().lower() if vals[2] else ''
                
                # Check if this looks like a segment code
                if re.match(r'^[A-Za-z0-9]+\.\d+\.\d+\.\d+', segment_code):
                    try:
                        length_val = float(vals[3]) if vals[3] else 0
                        qty_val = int(vals[4]) if vals[4] else 1
                        qty_per_product = int(vals[5]) if vals[5] else 1
                        
                        # Convert cm to mm
                        if unit == 'cm':
                            length_mm = length_val * 10
                        else:
                            length_mm = length_val  # Assume already mm
                        
                        # Normalize steel profile
                        steel_profile = normalize_steel_profile(steel_name)
                        
                        details.append({
                            'segment_name': segment_code,
                            'steel_profile': steel_profile,
                            'steel_raw': steel_name,  # Keep original for reference
                            'length_mm': length_mm,
                            'qty_per_unit': qty_val,
                            'qty_per_product': qty_per_product,
                            'sheet': sheet_name,
                            'file': os.path.basename(filepath)
                        })
                    except (ValueError, TypeError) as e:
                        pass  # Skip invalid rows
    
    return details

def import_all():
    """Main import function"""
    print("=" * 60)
    print("IMPORT ĐỊNH MỨC VẬT TƯ")
    print("=" * 60)
    
    # Phase 1: Steel Profiles
    print("\n[Phase 1] Creating Steel Profiles...")
    created_profiles = create_missing_steel_profiles()
    print(f"Created {len(created_profiles)} new profiles")
    
    # Phase 2: Parse all PhoiSat files
    print("\n[Phase 2] Parsing PhoiSat files...")
    all_details = []
    phoisat_files = glob.glob(f"{PHOISAT_DIR}/*.xlsx")
    
    for f in phoisat_files:
        try:
            details = parse_phoisat_file(f)
            all_details.extend(details)
            print(f"  ✓ {os.path.basename(f)}: {len(details)} segments")
        except Exception as e:
            print(f"  ✗ {os.path.basename(f)}: {e}")
    
    print(f"\nTotal segments parsed: {len(all_details)}")
    
    # Show sample
    print("\nSample data (first 5):")
    for d in all_details[:5]:
        print(f"  {d['segment_name']}: {d['steel_profile']} {d['length_mm']}mm x{d['qty_per_unit']}")
    
    return all_details

if __name__ == "__main__":
    # For testing outside bench
    all_details = []
    phoisat_files = glob.glob(f"{PHOISAT_DIR}/*.xlsx")
    
    for f in phoisat_files[:5]:  # Test first 5 files
        try:
            details = parse_phoisat_file(f)
            all_details.extend(details)
            print(f"✓ {os.path.basename(f)}: {len(details)} segments")
        except Exception as e:
            print(f"✗ {os.path.basename(f)}: {e}")
    
    print(f"\nTotal: {len(all_details)} segments")
    for d in all_details[:10]:
        print(f"  {d['segment_name']}: {d['steel_profile']} {d['length_mm']:.0f}mm x{d['qty_per_unit']}")
